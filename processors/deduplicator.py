"""
Deduplicator Module
Handles intelligence features and master Excel file management
FINAL HARD RESET: Fixed schema with positional mapping only
NO dictionary-based DataFrame creation - ONLY positional mapping
"""

import pandas as pd
import os
from datetime import datetime
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from processors.normalizer import normalize_string, normalize_amount, normalize_complaint_id


# FIXED COLUMN SCHEMA - Defined ONCE, used everywhere
COLUMNS = [
    "Complaint_ID",
    "Complaint_Date",
    "Incident_Date",
    "Category",
    "Sub_Category",
    "District",
    "State",
    "Amount_Lost",
    "Status",
    "Transaction_Count",
    "Data_Quality_Score",
    "Investigation_Ready",
    "Reporting_Delay_Days",
    "Reporting_Delay_Status",
    "Transaction_Pattern",
    "Source_File_Type"
]


def calculate_data_quality_score(complaint: Dict) -> int:
    """
    Feature 1: Complaint Normalization Score
    Add 20 points for each:
    - Complaint ID present
    - Complaint Date present
    - Amount Lost present
    - District + State present
    - At least one transaction present
    """
    score = 0
    
    complaint_id = str(complaint.get('Complaint_ID', '')).strip()
    if complaint_id and complaint_id != "Not Available":
        score += 20
    
    complaint_date = str(complaint.get('Complaint_Date', '')).strip()
    if complaint_date and complaint_date != "Not Available":
        score += 20
    
    amount = complaint.get('Amount_Lost', 0)
    if isinstance(amount, (int, float)) and amount > 0:
        score += 20
    
    district = str(complaint.get('District', '')).strip()
    state = str(complaint.get('State', '')).strip()
    if district and district != "Not Available" and state and state != "Not Available":
        score += 20
    
    transaction_count = complaint.get('Transaction_Count', 0)
    if isinstance(transaction_count, (int, float)) and transaction_count > 0:
        score += 20
    
    return score


def calculate_investigation_readiness(complaint: Dict) -> str:
    """
    Feature 2: Investigation Readiness Flag
    Set Investigation_Ready = YES if:
    - Amount Lost > 0
    - At least one transaction ID exists
    - Bank / Platform info exists
    Else: NO
    """
    amount = complaint.get('Amount_Lost', 0)
    amount_ok = isinstance(amount, (int, float)) and amount > 0
    
    transaction_count = complaint.get('Transaction_Count', 0)
    transaction_ok = isinstance(transaction_count, (int, float)) and transaction_count > 0
    
    bank_info = str(complaint.get('Bank_Platform_Info', '')).strip()
    bank_ok = bank_info and bank_info != "Not Available"
    
    if amount_ok and transaction_ok and bank_ok:
        return "YES"
    else:
        return "NO"


def calculate_reporting_delay(complaint: Dict) -> Tuple[int, str]:
    """
    Feature 3: Reporting Delay Indicator
    Compute Reporting_Delay_Days = Complaint_Date - Incident_Date
    If delay > 7 days: DELAYED, Else: ON_TIME
    Compute reporting delay ONLY if both dates exist
    If dates missing: reporting_delay_days = 0, reporting_delay_status = "NOT_AVAILABLE"
    """
    complaint_date_str = str(complaint.get('Complaint_Date', '')).strip()
    incident_date_str = str(complaint.get('Incident_Date', '')).strip()
    
    # Check if dates are missing or "Not Available"
    if not complaint_date_str or complaint_date_str == "Not Available":
        return 0, "NOT_AVAILABLE"
    
    if not incident_date_str or incident_date_str == "Not Available":
        return 0, "NOT_AVAILABLE"
    
    try:
        # Parse dates using pd.to_datetime (handles YYYY-MM-DD format)
        complaint_date = pd.to_datetime(complaint_date_str)
        incident_date = pd.to_datetime(incident_date_str)
        
        delay_days = (complaint_date - incident_date).days
        
        if delay_days > 7:
            status = "DELAYED"
        else:
            status = "ON_TIME"
        
        return delay_days, status
    except:
        return 0, "NOT_AVAILABLE"


def calculate_transaction_pattern(complaint: Dict) -> str:
    """
    Feature 4: Transaction Pattern Flag
    - One transaction AND amount > ₹50,000 → SINGLE_LARGE
    - Multiple transactions AND each < ₹10,000 → MULTIPLE_SMALL
    - Else → MIXED
    """
    transaction_count = complaint.get('Transaction_Count', 0)
    amount = complaint.get('Amount_Lost', 0)
    
    if not isinstance(transaction_count, (int, float)):
        transaction_count = 0
    if not isinstance(amount, (int, float)):
        amount = 0
    
    if transaction_count == 1 and amount > 50000:
        return "SINGLE_LARGE"
    elif transaction_count > 1:
        # Check if each transaction is small
        avg_per_transaction = amount / transaction_count if transaction_count > 0 else 0
        if avg_per_transaction < 10000:
            return "MULTIPLE_SMALL"
        else:
            return "MIXED"
    else:
        return "MIXED"


def apply_intelligence_features(complaints: List[Dict]) -> List[Dict]:
    """Apply all intelligence features to complaint data"""
    enhanced_complaints = []
    
    for complaint in complaints:
        enhanced = complaint.copy()
        
        # Feature 1: Data Quality Score
        enhanced['Data_Quality_Score'] = calculate_data_quality_score(complaint)
        
        # Feature 2: Investigation Readiness
        enhanced['Investigation_Ready'] = calculate_investigation_readiness(complaint)
        
        # Feature 3: Reporting Delay
        delay_days, delay_status = calculate_reporting_delay(complaint)
        enhanced['Reporting_Delay_Days'] = delay_days
        enhanced['Reporting_Delay_Status'] = delay_status
        
        # Feature 4: Transaction Pattern
        enhanced['Transaction_Pattern'] = calculate_transaction_pattern(complaint)
        
        enhanced_complaints.append(enhanced)
    
    return enhanced_complaints


def build_row_from_complaint(complaint: Dict, source_file_type: str) -> Dict:
    """
    Build a row dictionary with ONLY the fixed schema columns
    Uses explicit type casting as required
    NO extra keys, NO missing keys
    """
    # Extract and normalize all fields explicitly
    complaint_id_raw = complaint.get('Complaint_ID', '')
    complaint_id = normalize_complaint_id(complaint_id_raw)
    
    complaint_date_raw = complaint.get('Complaint_Date', '')
    complaint_date = normalize_string(complaint_date_raw)
    
    incident_date_raw = complaint.get('Incident_Date', '')
    incident_date = normalize_string(incident_date_raw)
    
    category_raw = complaint.get('Category', '')
    category = normalize_string(category_raw)
    
    sub_category_raw = complaint.get('Sub_Category', '')
    sub_category = normalize_string(sub_category_raw)
    
    district_raw = complaint.get('District', '')
    district = normalize_string(district_raw)
    
    state_raw = complaint.get('State', '')
    state = normalize_string(state_raw)
    
    amount_lost_raw = complaint.get('Amount_Lost', 0)
    amount_lost = normalize_amount(amount_lost_raw)
    
    status_raw = complaint.get('Status', 'Pending')
    status = normalize_string(status_raw)
    
    transaction_count_raw = complaint.get('Transaction_Count', 0)
    transaction_count = int(transaction_count_raw) if transaction_count_raw else 0
    
    data_quality_score_raw = complaint.get('Data_Quality_Score', 0)
    data_quality_score = int(data_quality_score_raw) if data_quality_score_raw else 0
    
    investigation_ready_raw = complaint.get('Investigation_Ready', 'NO')
    investigation_ready = str(investigation_ready_raw)
    
    reporting_delay_days_raw = complaint.get('Reporting_Delay_Days', 0)
    reporting_delay_days = int(reporting_delay_days_raw) if reporting_delay_days_raw else 0
    
    reporting_delay_status_raw = complaint.get('Reporting_Delay_Status', 'UNKNOWN')
    reporting_delay_status = str(reporting_delay_status_raw)
    
    transaction_pattern_raw = complaint.get('Transaction_Pattern', 'MIXED')
    transaction_pattern = str(transaction_pattern_raw)
    
    source_file_type_str = str(source_file_type).upper()
    
    # Build row with EXACT schema - explicit type casting
    row = {
        "Complaint_ID": str(complaint_id),  # MUST be NCRP acknowledgement number
        "Complaint_Date": str(complaint_date),
        "Incident_Date": str(incident_date),
        "Category": str(category),
        "Sub_Category": str(sub_category),
        "District": str(district),
        "State": str(state),
        "Amount_Lost": float(amount_lost),
        "Status": str(status),
        "Transaction_Count": int(transaction_count),
        "Data_Quality_Score": int(data_quality_score),
        "Investigation_Ready": str(investigation_ready),
        "Reporting_Delay_Days": int(reporting_delay_days),
        "Reporting_Delay_Status": str(reporting_delay_status),
        "Transaction_Pattern": str(transaction_pattern),
        "Source_File_Type": str(source_file_type_str)
    }
    
    return row


def safe_write_excel(df: pd.DataFrame, filepath: str):
    """
    Safely write DataFrame to Excel with explicit type formatting
    Prevents scientific notation and type inference issues
    Sets column widths and formats headers
    """
    # Ensure DataFrame columns match fixed schema exactly
    if list(df.columns) != COLUMNS:
        # Create new DataFrame with fixed schema
        data = []
        for idx in range(len(df)):
            row_data = []
            for col in COLUMNS:
                if col in df.columns:
                    row_data.append(df.iloc[idx][col])
                else:
                    # Default values for missing columns
                    if col == 'Amount_Lost':
                        row_data.append(0.0)
                    elif col in ['Transaction_Count', 'Data_Quality_Score', 'Reporting_Delay_Days']:
                        row_data.append(0)
                    else:
                        row_data.append('Not Available')
            data.append(row_data)
        df = pd.DataFrame(data, columns=COLUMNS)
    
    # Force data types BEFORE writing
    df['Complaint_ID'] = df['Complaint_ID'].astype(str)
    df['Complaint_Date'] = df['Complaint_Date'].astype(str)
    df['Incident_Date'] = df['Incident_Date'].astype(str)
    df['Category'] = df['Category'].astype(str)
    df['Sub_Category'] = df['Sub_Category'].astype(str)
    df['District'] = df['District'].astype(str)
    df['State'] = df['State'].astype(str)
    df['Status'] = df['Status'].astype(str)
    df['Investigation_Ready'] = df['Investigation_Ready'].astype(str)
    df['Reporting_Delay_Status'] = df['Reporting_Delay_Status'].astype(str)
    df['Transaction_Pattern'] = df['Transaction_Pattern'].astype(str)
    df['Source_File_Type'] = df['Source_File_Type'].astype(str)
    
    df['Amount_Lost'] = pd.to_numeric(df['Amount_Lost'], errors='coerce').fillna(0.0).astype(float)
    df['Transaction_Count'] = pd.to_numeric(df['Transaction_Count'], errors='coerce').fillna(0).astype(int)
    df['Data_Quality_Score'] = pd.to_numeric(df['Data_Quality_Score'], errors='coerce').fillna(0).astype(int)
    df['Reporting_Delay_Days'] = pd.to_numeric(df['Reporting_Delay_Days'], errors='coerce').fillna(0).astype(int)
    
    # Write using pandas with openpyxl engine
    df.to_excel(filepath, index=False, engine='openpyxl')
    
    # Now open with openpyxl to apply explicit formatting
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Set column widths to prevent ########
    column_widths = {
        'A': 25,  # Complaint_ID
        'B': 15,  # Complaint_Date
        'C': 15,  # Incident_Date
        'D': 20,  # Category
        'E': 30,  # Sub_Category
        'F': 20,  # District
        'G': 20,  # State
        'H': 18,  # Amount_Lost
        'I': 15,  # Status
        'J': 18,  # Transaction_Count
        'K': 20,  # Data_Quality_Score
        'L': 22,  # Investigation_Ready
        'M': 22,  # Reporting_Delay_Days
        'N': 25,  # Reporting_Delay_Status
        'O': 25,  # Transaction_Pattern
        'P': 18   # Source_File_Type
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Apply formatting to each column
    header_row = 1
    for col_idx, header in enumerate(COLUMNS, start=1):
        col_letter = ws.cell(row=header_row, column=col_idx).column_letter
        
        # Format header
        header_cell = ws.cell(row=header_row, column=col_idx)
        header_cell.font = Font(bold=True)
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply text format to Complaint_ID column
        if header == 'Complaint_ID':
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.value = str(cell.value)
                cell.number_format = '@'  # Text format
        
        # Format Amount_Lost column to prevent scientific notation
        elif header == 'Amount_Lost':
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    try:
                        amount = float(cell.value)
                        cell.value = amount
                        cell.number_format = '#,##0.00'  # Number format with commas
                    except:
                        pass
    
    # Save the workbook
    wb.save(filepath)


def append_to_master_excel(complaints: List[Dict], source_file_type: str) -> Tuple[int, int]:
    """
    Append complaints to master Excel file with deduplication
    Uses FIXED SCHEMA + POSITIONAL MAPPING ONLY
    NO dictionary-based DataFrame creation
    Returns: (new_count, total_count)
    """
    master_file = 'output/ncrp_master.xlsx'
    
    # Apply intelligence features
    enhanced_complaints = apply_intelligence_features(complaints)
    
    # Build rows using fixed schema - explicit extraction
    new_rows = []
    for complaint in enhanced_complaints:
        row = build_row_from_complaint(complaint, source_file_type)
        new_rows.append(row)
    
    # Create DataFrame using POSITIONAL MAPPING ONLY
    # This is the MANDATORY approach - no dictionary-based creation
    if new_rows:
        data = [[row[col] for col in COLUMNS] for row in new_rows]
        new_df = pd.DataFrame(data, columns=COLUMNS)
    else:
        new_df = pd.DataFrame(columns=COLUMNS)
    
    # Load existing master file if it exists
    existing_df = pd.DataFrame(columns=COLUMNS)
    if os.path.exists(master_file):
        try:
            # Read existing file
            existing_raw = pd.read_excel(master_file, dtype={'Complaint_ID': str})
            
            # Rebuild using fixed schema - positional mapping
            if not existing_raw.empty:
                existing_data = []
                for idx in range(len(existing_raw)):
                    row_data = []
                    for col in COLUMNS:
                        if col in existing_raw.columns:
                            val = existing_raw.iloc[idx][col]
                            # Type conversion based on column
                            if col == 'Complaint_ID':
                                row_data.append(str(val))
                            elif col == 'Amount_Lost':
                                row_data.append(float(val) if pd.notna(val) else 0.0)
                            elif col in ['Transaction_Count', 'Data_Quality_Score', 'Reporting_Delay_Days']:
                                row_data.append(int(val) if pd.notna(val) else 0)
                            else:
                                row_data.append(str(val) if pd.notna(val) else 'Not Available')
                        else:
                            # Default values for missing columns
                            if col == 'Amount_Lost':
                                row_data.append(0.0)
                            elif col in ['Transaction_Count', 'Data_Quality_Score', 'Reporting_Delay_Days']:
                                row_data.append(0)
                            else:
                                row_data.append('Not Available')
                    existing_data.append(row_data)
                
                existing_df = pd.DataFrame(existing_data, columns=COLUMNS)
        except Exception as e:
            # If reading fails, start fresh
            existing_df = pd.DataFrame(columns=COLUMNS)
    
    # Get existing Complaint_IDs for deduplication
    existing_ids = set()
    if not existing_df.empty:
        existing_ids = set(existing_df['Complaint_ID'].astype(str).str.strip())
    
    # Filter out duplicates from new rows
    filtered_rows = []
    for row in new_rows:
        complaint_id = str(row['Complaint_ID']).strip()
        # Skip if empty or "Not Available" (but allow if it's a generated ID)
        if complaint_id and complaint_id != "Not Available" and complaint_id not in existing_ids:
            filtered_rows.append(row)
            existing_ids.add(complaint_id)
    
    new_count = len(filtered_rows)
    
    # Combine existing and new records using POSITIONAL MAPPING
    if filtered_rows:
        # Build DataFrame with explicit schema - positional mapping
        filtered_data = [[row[col] for col in COLUMNS] for row in filtered_rows]
        filtered_df = pd.DataFrame(filtered_data, columns=COLUMNS)
        
        if existing_df.empty:
            combined_df = filtered_df
        else:
            # Concatenate and rebuild to ensure correct schema
            combined_temp = pd.concat([existing_df, filtered_df], ignore_index=True)
            # Rebuild using positional mapping to ensure correctness
            combined_data = [[combined_temp.iloc[idx][col] for col in COLUMNS] for idx in range(len(combined_temp))]
            combined_df = pd.DataFrame(combined_data, columns=COLUMNS)
    else:
        combined_df = existing_df
    
    # Ensure we have a DataFrame (even if empty)
    if combined_df.empty:
        combined_df = pd.DataFrame(columns=COLUMNS)
    
    # Force data types one final time before saving
    combined_df['Complaint_ID'] = combined_df['Complaint_ID'].astype(str)
    combined_df['Complaint_Date'] = combined_df['Complaint_Date'].astype(str)
    combined_df['Incident_Date'] = combined_df['Incident_Date'].astype(str)
    combined_df['Category'] = combined_df['Category'].astype(str)
    combined_df['Sub_Category'] = combined_df['Sub_Category'].astype(str)
    combined_df['District'] = combined_df['District'].astype(str)
    combined_df['State'] = combined_df['State'].astype(str)
    combined_df['Status'] = combined_df['Status'].astype(str)
    combined_df['Investigation_Ready'] = combined_df['Investigation_Ready'].astype(str)
    combined_df['Reporting_Delay_Status'] = combined_df['Reporting_Delay_Status'].astype(str)
    combined_df['Transaction_Pattern'] = combined_df['Transaction_Pattern'].astype(str)
    combined_df['Source_File_Type'] = combined_df['Source_File_Type'].astype(str)
    
    combined_df['Amount_Lost'] = pd.to_numeric(combined_df['Amount_Lost'], errors='coerce').fillna(0.0).astype(float)
    combined_df['Transaction_Count'] = pd.to_numeric(combined_df['Transaction_Count'], errors='coerce').fillna(0).astype(int)
    combined_df['Data_Quality_Score'] = pd.to_numeric(combined_df['Data_Quality_Score'], errors='coerce').fillna(0).astype(int)
    combined_df['Reporting_Delay_Days'] = pd.to_numeric(combined_df['Reporting_Delay_Days'], errors='coerce').fillna(0).astype(int)
    
    # Save to master Excel file using safe writing
    os.makedirs('output', exist_ok=True)
    safe_write_excel(combined_df, master_file)
    
    total_count = len(combined_df)
    
    return new_count, total_count