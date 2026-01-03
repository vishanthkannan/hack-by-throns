"""
NCRP Complaint Automation Tool - Unified System
Handles PDF, CSV, and Excel files with Excel export
"""

from flask import Flask, request, render_template, jsonify, send_file
from werkzeug.utils import secure_filename
import pdfplumber
import pandas as pd
import re
import os
from typing import Dict, List, Optional

app = Flask(__name__, template_folder='viewer_templates')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'csv', 'xlsx', 'xls'}
app.config['MASTER_EXCEL'] = 'ncrp_master.xlsx'

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('output', exist_ok=True)


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# ==================== PDF EXTRACTION (KEEP EXISTING LOGIC) ====================

def extract_text_from_pdf(filepath):
    """Extract all text from PDF file"""
    text = ""
    try:
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
    except Exception as e:
        raise Exception(f"Error reading PDF: {str(e)}")
    return text


def normalize_text(text):
    """Normalize PDF text"""
    text = text.replace("\r", "\n")
    text = re.sub(r"\n+", "\n", text)
    return text


def extract_field(pattern, text):
    """Extract field using regex pattern (layout-agnostic)"""
    try:
        match = re.search(pattern, text, re.IGNORECASE)
        if match and match.groups():
            result = match.group(1).strip()
            result = result.replace('\n', ' ').replace('\r', ' ')
            result = re.sub(r'\s+', ' ', result).strip()
            stop_patterns = [r'\b(District|State|Category|Sub|Complaint|Incident|Amount|Total)\b']
            for stop_pattern in stop_patterns:
                stop_match = re.search(stop_pattern, result, re.IGNORECASE)
                if stop_match:
                    result = result[:stop_match.start()].strip()
                    break
            return result if result else None
    except:
        pass
    return None


def extract_from_pdf(filepath) -> Optional[Dict]:
    """Extract complaint data from PDF using existing working logic

    Adds extraction of the 'Action Taken / Remarks' area from the last page and
    derives a readable `Status` using a simple rule-based mapper (no ML/NLP).
    """
    try:
        text = extract_text_from_pdf(filepath)
        
        if not text or len(text.strip()) < 50:
            return None
        
        text = normalize_text(text)
        
        # Extract fields using layout-agnostic regex patterns
        complaint_id = extract_field(r"Acknowledgement\s*Number\s*:\s*(\d+)", text)
        complaint_date = extract_field(r"Complaint\s+Date\s*[:\-]?\s*([0-9/]+)", text)
        incident_date = extract_field(r"Incident\s+Date/Time\s*[:\-]?\s*([0-9/]+)", text)
        category = extract_field(r"Category\s+of\s+complaint\s*[:\-]?\s*([A-Za-z ]+)", text)
        sub_category = extract_field(r"Sub\s+Category\s+of\s+Complaint\s*[:\-]?\s*([A-Za-z ]+)", text)
        district = extract_field(r"District\s*[:\-]?\s*([A-Za-z ]+)", text)
        state = extract_field(r"State\s*[:\-]?\s*([A-Za-z ]+)", text)
        amount_lost = extract_field(r"Total\s+Fraudulent\s+Amount\s+reported\s+by\s+complainant\s*:\s*([0-9,\.]+)", text)
        
        # Helper: extract Action Taken / Remarks from the LAST PAGE of the PDF
        def extract_action_taken_remarks_from_pdf(path: str) -> str:
            try:
                with pdfplumber.open(path) as pdf:
                    last_page = pdf.pages[-1]
                    last_text = last_page.extract_text() or ''
                    last_text = normalize_text(last_text)
                    # Common markers seen in NCRP PDFs
                    markers = [r"Action\s*Taken\s*/?\s*Remarks", r"Action\s*Taken", r"Remarks"]
                    for marker in markers:
                        m = re.search(marker + r"\s*[:\-]?\s*(.*)", last_text, re.IGNORECASE | re.DOTALL)
                        if m:
                            captured = m.group(1).strip()
                            # Stop at common next-label patterns (lines like "Label:")
                            captured = re.split(r"\n[A-Za-z][A-Za-z\s]{1,40}\s*:", captured)[0]
                            captured = captured.replace('\n', ' ').strip()
                            captured = re.sub(r'\s+', ' ', captured)
                            return captured
            except Exception:
                pass
            return ""
        
        # Helper: rule-based status mapper (strictly rules per requirements)
        def map_status_from_action(action_text: str) -> str:
            if not action_text or not action_text.strip():
                # If nothing is present in Action Taken, default to Registered per rules
                return "Registered"
            s = action_text.lower()
            # Check specific phrases in order of priority
            if "under enquiry" in s or "under investigation" in s:
                return "Under Enquiry"
            if "fir registered" in s or "fir filed" in s:
                return "FIR Filed"
            if "closed" in s or "disposed" in s:
                return "Closed"
            if "frozen" in s or "amount blocked" in s:
                return "Amount Frozen"
            if "forwarded" in s or "assigned" in s:
                return "Registered"
            # Default fallback
            return "Registered"
        
        # Format amount for display
        if amount_lost:
            try:
                amount_clean = amount_lost.replace(',', '')
                amount_float = float(amount_clean)
                amount_lost = f"{amount_float:,.2f}"
            except:
                pass
        
        # Extract Action Taken / Remarks from last page (if present)
        action_taken = extract_action_taken_remarks_from_pdf(filepath)
        derived_status = map_status_from_action(action_taken)
        
        return {
            'Complaint_ID': complaint_id if complaint_id else "Not Available",
            'Complaint_Date': complaint_date if complaint_date else "Not Available",
            'Incident_Date': incident_date if incident_date else "Not Available",
            'Category': category if category else "Not Available",
            'Sub_Category': sub_category if sub_category else "Not Available",
            'District': district if district else "Not Available",
            'State': state if state else "Not Available",
            'Amount_Lost': amount_lost if amount_lost else "Not Available",
            'Status': derived_status if derived_status else "Registered",
            'Action_Taken_Remarks': action_taken if action_taken else "Not Available",
            'Source_File_Type': 'PDF'
        }
        
    except Exception as e:
        raise Exception(f"Error extracting PDF data: {str(e)}")


# ==================== CSV/EXCEL PROCESSING ====================

def normalize_column_name(col: str) -> str:
    """Normalize column names to standard format"""
    col = str(col).strip().lower()
    mappings = {
        'complaint id': 'Complaint_ID',
        'acknowledgement number': 'Complaint_ID',
        'complaint number': 'Complaint_ID',
        'complaint date': 'Complaint_Date',
        'incident date': 'Incident_Date',
        'category': 'Category',
        'sub category': 'Sub_Category',
        'subcategory': 'Sub_Category',
        'district': 'District',
        'state': 'State',
        'amount': 'Amount_Lost',
        'amount lost': 'Amount_Lost',
        'fraudulent amount': 'Amount_Lost',
        'status': 'Status',
        # Action Taken / Remarks mappings (various forms)
        'action taken': 'Action_Taken_Remarks',
        'action taken / remarks': 'Action_Taken_Remarks',
        'action taken/remarks': 'Action_Taken_Remarks',
        'action_taken_remarks': 'Action_Taken_Remarks',
        'remarks': 'Action_Taken_Remarks',
    }
    return mappings.get(col, col)


def normalize_value(value) -> str:
    """Normalize value to string, handle NaN"""
    if pd.isna(value) or value == '' or value is None:
        return "Not Available"
    return str(value).strip()


def extract_from_csv(filepath) -> List[Dict]:
    """Extract complaint data from CSV file"""
    try:
        encodings = ['utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(filepath, encoding=encoding)
                break
            except:
                continue
        
        if df is None or df.empty:
            return []
        
        # Normalize column names
        df.columns = [normalize_column_name(col) for col in df.columns]
        
        complaints = []
        for idx, row in df.iterrows():
            complaint = {
                'Complaint_ID': normalize_value(row.get('Complaint_ID', '')),
                'Complaint_Date': normalize_value(row.get('Complaint_Date', '')),
                'Incident_Date': normalize_value(row.get('Incident_Date', '')),
                'Category': normalize_value(row.get('Category', '')),
                'Sub_Category': normalize_value(row.get('Sub_Category', '')),
                'District': normalize_value(row.get('District', '')),
                'State': normalize_value(row.get('State', '')),
                'Amount_Lost': normalize_value(row.get('Amount_Lost', '')),
                'Status': normalize_value(row.get('Status', 'Pending')),
                'Action_Taken_Remarks': normalize_value(row.get('Action_Taken_Remarks', '')),
                'Source_File_Type': 'CSV'
            }
            complaints.append(complaint)
        
        return complaints
        
    except Exception as e:
        raise Exception(f"Error extracting CSV data: {str(e)}")


def extract_from_excel(filepath) -> List[Dict]:
    """Extract complaint data from Excel file"""
    try:
        df = pd.read_excel(filepath, sheet_name=0)
        
        if df.empty:
            return []
        
        # Normalize column names
        df.columns = [normalize_column_name(col) for col in df.columns]
        
        complaints = []
        for idx, row in df.iterrows():
            complaint = {
                'Complaint_ID': normalize_value(row.get('Complaint_ID', '')),
                'Complaint_Date': normalize_value(row.get('Complaint_Date', '')),
                'Incident_Date': normalize_value(row.get('Incident_Date', '')),
                'Category': normalize_value(row.get('Category', '')),
                'Sub_Category': normalize_value(row.get('Sub_Category', '')),
                'District': normalize_value(row.get('District', '')),
                'State': normalize_value(row.get('State', '')),
                'Amount_Lost': normalize_value(row.get('Amount_Lost', '')),
                'Status': normalize_value(row.get('Status', 'Pending')),
                'Action_Taken_Remarks': normalize_value(row.get('Action_Taken_Remarks', '')),
                'Source_File_Type': 'EXCEL'
            }
            complaints.append(complaint)
        
        return complaints
        
    except Exception as e:
        raise Exception(f"Error extracting Excel data: {str(e)}")


# ==================== EXCEL EXPORT WITH DUPLICATE CHECK ====================

def save_to_master_excel(complaints: List[Dict]) -> Dict:
    """Save complaints to master Excel with duplicate check"""
    master_file = os.path.join('output', app.config['MASTER_EXCEL'])
    
    # Define columns
    columns = [
        'Complaint_ID',
        'Complaint_Date',
        'Incident_Date',
        'Category',
        'Sub_Category',
        'District',
        'State',
        'Amount_Lost',
        'Status',
        'Action_Taken_Remarks',
        'Source_File_Type'
    ]
    
    # --- After writing combined_df to Excel, apply openpyxl formatting for readability ---
    def _format_master_excel(path: str):
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            wb = load_workbook(path)
            ws = wb.active

            # Header formatting: bold, center, wrap text
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', wrap_text=True)

            # Column widths mapping
            col_widths = {
                'Complaint_ID': 22,
                'Complaint_Date': 15,
                'Incident_Date': 15,
                'Category': 25,
                'Sub_Category': 20,
                'District': 18,
                'State': 15,
                'Amount_Lost': 15,
                'Status': 18,
                'Action_Taken_Remarks': 40,
                'Source_File_Type': 15,
            }

            # Map headers to column index
            header_index = {ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)}

            # Apply widths
            for col_name, width in col_widths.items():
                col_idx = header_index.get(col_name)
                if col_idx:
                    letter = get_column_letter(col_idx)
                    ws.column_dimensions[letter].width = width

            # Data rows: vertical center and wrap text
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Ensure Complaint_ID stored as TEXT to avoid scientific notation
            if 'Complaint_ID' in header_index:
                col = header_index['Complaint_ID']
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, max_row=ws.max_row):
                    c = row[0]
                    if c.value is None:
                        continue
                    c.value = str(c.value)
                    c.number_format = '@'

            # Format Amount_Lost as numeric with two decimal places where possible
            if 'Amount_Lost' in header_index:
                col = header_index['Amount_Lost']
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col, max_row=ws.max_row):
                    c = row[0]
                    if c.value is None:
                        continue
                    v = str(c.value).replace(',', '').strip()
                    try:
                        num = float(v)
                        c.value = num
                        c.number_format = '#,##0.00'
                    except:
                        # Leave as-is if cannot parse (e.g., 'Not Available')
                        pass

            wb.save(path)
        except Exception:
            # Do not raise - formatting is best-effort and should not break saving
            pass
    
    # Read existing file if it exists
    existing_df = pd.DataFrame(columns=columns)
    if os.path.exists(master_file):
        try:
            existing_df = pd.read_excel(master_file, dtype={'Complaint_ID': str})
            # Ensure columns match
            for col in columns:
                if col not in existing_df.columns:
                    existing_df[col] = 'Not Available'
            existing_df = existing_df[columns]
        except:
            existing_df = pd.DataFrame(columns=columns)
    
    # Get existing Complaint_IDs (as strings)
    existing_ids = set()
    if not existing_df.empty:
        existing_ids = set(existing_df['Complaint_ID'].astype(str).str.strip())
    
    # Filter out duplicates
    new_complaints = []
    duplicates = []
    
    for complaint in complaints:
        complaint_id = str(complaint.get('Complaint_ID', '')).strip()
        if complaint_id and complaint_id != "Not Available":
            if complaint_id not in existing_ids:
                new_complaints.append(complaint)
                existing_ids.add(complaint_id)
            else:
                duplicates.append(complaint_id)
        else:
            # Skip complaints without valid ID
            continue
    
    # Combine existing and new
    if new_complaints:
        new_df = pd.DataFrame(new_complaints)
        # Ensure Complaint_ID is string
        new_df['Complaint_ID'] = new_df['Complaint_ID'].astype(str)
        
        if existing_df.empty:
            combined_df = new_df
        else:
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    else:
        combined_df = existing_df
    
    # Ensure all columns exist and in correct order
    for col in columns:
        if col not in combined_df.columns:
            combined_df[col] = 'Not Available'
    combined_df = combined_df[columns]
    
    # Ensure Complaint_ID is string
    combined_df['Complaint_ID'] = combined_df['Complaint_ID'].astype(str)
    
    # Save to Excel
    combined_df.to_excel(master_file, index=False, engine='openpyxl')

    # Apply formatting (best-effort) to make the master Excel readable
    try:
        _format_master_excel(master_file)
    except Exception:
        pass
    
    return {
        'new_count': len(new_complaints),
        'duplicate_count': len(duplicates),
        'total_count': len(combined_df),
        'duplicates': duplicates
    }


# ==================== FLASK ROUTES ====================

@app.route('/')
def index():
    """Render main page"""
    return render_template('viewer_index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and extraction"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'message': 'Only PDF, CSV, and Excel files are allowed'}), 400
        
        # Save uploaded file
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower()
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Extract data based on file type
        complaints = []
        if file_ext == 'pdf':
            complaint = extract_from_pdf(filepath)
            if complaint:
                complaints = [complaint]
        elif file_ext == 'csv':
            complaints = extract_from_csv(filepath)
        elif file_ext in ['xlsx', 'xls']:
            complaints = extract_from_excel(filepath)
        
        # Clean up uploaded file
        try:
            os.remove(filepath)
        except:
            pass
        
        if not complaints:
            return jsonify({
                'success': False,
                'message': 'Could not extract data from file'
            }), 400
        
        return jsonify({
            'success': True,
            'data': complaints,  # List of complaints (PDF=1, CSV/Excel=multiple)
            'count': len(complaints)
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        }), 500


@app.route('/save', methods=['POST'])
def save_to_excel():
    """Save extracted complaints to master Excel"""
    try:
        data = request.get_json()
        complaints = data.get('complaints', [])
        
        if not complaints:
            return jsonify({
                'success': False,
                'message': 'No complaints to save'
            }), 400
        
        result = save_to_master_excel(complaints)
        
        if result['new_count'] > 0:
            message = f"Successfully saved {result['new_count']} new complaint(s). Total: {result['total_count']}"
            if result['duplicate_count'] > 0:
                message += f". {result['duplicate_count']} duplicate(s) skipped."
        else:
            message = f"No new complaints saved. {result['duplicate_count']} duplicate(s) found."
        
        return jsonify({
            'success': True,
            'message': message,
            'new_count': result['new_count'],
            'duplicate_count': result['duplicate_count'],
            'total_count': result['total_count']
        })
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error saving to Excel: {str(e)}'
        }), 500


@app.route('/download', methods=['GET'])
def download_excel():
    """Download the master Excel file"""
    try:
        master_file = os.path.join('output', app.config['MASTER_EXCEL'])
        
        # Check if file exists
        if not os.path.exists(master_file):
            return jsonify({
                'success': False,
                'message': 'Master Excel file not found. Please save a complaint first.'
            }), 404
        
        # Send file securely
        return send_file(
            master_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=app.config['MASTER_EXCEL']
        )
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error downloading file: {str(e)}'
        }), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
